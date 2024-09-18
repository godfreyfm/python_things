import openpyxl  # External package that works with spreadsheets specifically. Make sure to install
from openpyxl.styles import Font


def load_inventory(file_path):
    """Loads the inventory from an Excel file."""
    workbook = openpyxl.load_workbook(file_path)
    return workbook["Sheet1"], workbook


def process_inventory(product_list):
    """Processes inventory to calculate products per supplier, total value per supplier, and low inventory items."""
    products_per_supplier = {}
    total_value_per_supplier = {}
    products_under_10_inv = {}

    for product_row in range(2, product_list.max_row + 1):
        supplier_name = product_list.cell(product_row, 4).value
        inventory = product_list.cell(product_row, 2).value
        price = product_list.cell(product_row, 3).value
        product_num = product_list.cell(product_row, 1).value

        # Update product count per supplier
        update_products_per_supplier(products_per_supplier, supplier_name)

        # Update total inventory value per supplier
        update_total_value_per_supplier(total_value_per_supplier, supplier_name, inventory, price)

        # Check for products with inventory less than 10
        if inventory < 10:
            products_under_10_inv[int(product_num)] = int(inventory)

        # Calculate total inventory price
        update_inventory_price(product_list, product_row, inventory, price)

    return products_per_supplier, total_value_per_supplier, products_under_10_inv


def update_products_per_supplier(products_per_supplier, supplier_name):
    """Updates the count of products per supplier."""
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] += 1
    else:
        products_per_supplier[supplier_name] = 1


def update_total_value_per_supplier(total_value_per_supplier, supplier_name, inventory, price):
    """Calculates and updates the total inventory value per supplier."""
    total_value = inventory * price
    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] += total_value
    else:
        total_value_per_supplier[supplier_name] = total_value


def update_inventory_price(product_list, product_row, inventory, price):
    """Updates the cell with the total inventory price."""
    inventory_price_cell = product_list.cell(product_row, 5)  # Assuming this is the 5th column for total price
    inventory_price_cell.value = inventory * price


def update_excel_with_total_header(product_list):
    """Updates the Excel sheet with the total header and formatting."""
    total_header_cell = product_list.cell(column=5, row=1)
    total_header_cell.value = "Total"
    total_header_cell.font = Font(size=11, bold=True)


def main():
    """Main function to run the inventory processing."""
    product_list, inv_file = load_inventory("inventory.xlsx")
    products_per_supplier, total_value_per_supplier, products_under_10_inv = process_inventory(product_list)

    print(products_per_supplier)
    print(total_value_per_supplier)
    print(products_under_10_inv)

    update_excel_with_total_header(product_list)

    inv_file.save("inventory_with_total_value.xlsx")


if __name__ == "__main__":
    main()
